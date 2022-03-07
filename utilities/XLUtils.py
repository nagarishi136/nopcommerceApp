import openpyxl
def getRowCount(file,sheatName):
    workbook=openpyxl.load_workbook(file)
    sheat=workbook[sheatName]
    return (sheat.max_row)
def getColumnCount(file,sheatName):
    workbook=openpyxl.load_workbook(file)
    sheat=workbook[sheatName]
    return (sheat.max_column)
def readData(file,shetaName,rownum,columnno):
    workbook=openpyxl.load_workbook(file)
    sheat=workbook[shetaName]
    return sheat.cell(row=rownum,column=columnno).value
def writeData(file,sheatName,rownum,columnno,data):
    workbook=openpyxl.load_workbook(file)
    sheat=workbook[sheatName]
    sheat.cell(row=rownum,column=columnno).value=data
    workbook.save(file)
